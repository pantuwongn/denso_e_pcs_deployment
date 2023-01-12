import os
import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side

from e_pcs_drawer import (
    PCSSheetDrawer
)
from e_pcs_form_page import (
    PCSPageBuilder,
    PCSHeaderPage,
    PCSItemPage,
)

class PCSForm:
    def __init__(self, templatePath: str, dataDict: dict):
        self.templatePath = templatePath
        self.templateSheetName = 'empty'
        self.outputDir = 'output'
        self.tempDir = 'temp'
        self.maxItemHeight = 51
        self.itemStartRow = 12

        self.dataDict = dataDict
        self._initializeTemplateWorkbook()

    def _initializeTemplateWorkbook(self):
        self.workbook = load_workbook(filename = self.templatePath)

    def generate(self, fileName: str):
        self._createTempDir()

        pageBuilder = PCSPageBuilder(self.maxItemHeight, self.dataDict)
        pcsHeaderPageList = pageBuilder.create()
        
        for i, pcsHeaderPage in enumerate(pcsHeaderPageList):
            templateSheet = self.workbook[self.templateSheetName]
            pageSheet = self.workbook.copy_worksheet(templateSheet)
            drawer = PCSSheetDrawer(self.maxItemHeight, self.itemStartRow, self.tempDir, pageSheet)

            self._setSheetTitle(pageSheet, pcsHeaderPage, i + 1)
            self._writeHeaders(drawer, pcsHeaderPage)
            self._mergeHeaders(pageSheet)
            self._mergeBody(pageSheet)

            drawer.drawCheckProcessDashedLine()

            currentRow = self.itemStartRow

            for j, pcsItemPage in enumerate(pcsHeaderPage.itemList):
                #   column marking
                numberCol = 4
                parameterCol = 5
                controlMethodIntervalCol = 9
                controlMethodCol = 10
                controlMethodControlPersonCol = 11
                processCapabilityCol = 12
                remarkCol = 13
                relateStandardCol = 15

                #   no col
                drawer.writeSubBody(currentRow, numberCol, pcsItemPage.controlItemNo)

                #   parameter col
                drawer.writeBody(currentRow, parameterCol, pcsItemPage.parameterDetail, 'left')
                drawer.writeBody(currentRow + len(pcsItemPage.parameterDetail), parameterCol, pcsItemPage.measuredValue, 'left')
                drawer.writeBody(currentRow + pcsItemPage.height - 1, parameterCol, pcsItemPage.measurement, 'left')

                #   control method interval col
                drawer.writeBody(currentRow, controlMethodIntervalCol, pcsItemPage.controlMethodIntervalType)
                drawer.writeBody(currentRow + len(pcsItemPage.controlMethodIntervalType), controlMethodIntervalCol, pcsItemPage.controlMethodIntervalPeriod)
                drawer.writeBody(currentRow + len(pcsItemPage.controlMethodIntervalType) + len(pcsItemPage.controlMethodIntervalPeriod), controlMethodIntervalCol, pcsItemPage.controlMethodSampleNo)
                drawer.writeBody(currentRow + pcsItemPage.height - 1, controlMethodIntervalCol, pcsItemPage.controlMethodCalibrationInterval)

                #   control method col
                drawer.writeBody(currentRow, controlMethodCol, pcsItemPage.controlMethodType)
                drawer.writeBody(currentRow + len(pcsItemPage.controlMethodType), controlMethodCol, pcsItemPage.controlMethodItemType)
                drawer.writeBody(currentRow + pcsItemPage.height - 1, controlMethodCol, pcsItemPage.controlMethod)

                #   control method control person col
                drawer.writeBody(currentRow, controlMethodControlPersonCol, pcsItemPage.controlMethodControlPerson)

                #   process capability col
                drawer.writeBody(currentRow, processCapabilityCol, pcsItemPage.processCapability)

                #   remark col
                drawer.writeBody(currentRow, remarkCol, pcsItemPage.remark, 'left')

                #   relate standard col
                drawer.writeBody(currentRow, relateStandardCol, pcsItemPage.relateStandard)

                #   create border
                self._createItemBorder(pageSheet, currentRow, pcsItemPage)

                #   draw sc symbol
                drawer.drawScSymbolList(currentRow, pcsItemPage.height, pcsItemPage.scSymbolList)

                #   draw control item
                drawer.drawControlItemSymbol(pcsItemPage.controlItemType, currentRow, pcsItemPage.height)

                currentRow += pcsItemPage.height

            #   draw check timing
            drawer.drawCheckTiming(pcsHeaderPage.itemList, pcsHeaderPage.summaryTimingDict)

            #   draw control item connector
            beforePageItem = pcsHeaderPageList[i - 1].itemList[-1] if i - 1 >= 0 else None
            afterPageItem = pcsHeaderPageList[i + 1].itemList[0] if i + 1 < len(pcsHeaderPageList) else None
            drawer.drawControlItemConnectorGroup(pcsHeaderPage.itemList, pcsHeaderPage.summaryTimingDict, beforePageItem, afterPageItem)
        
        self._saveWorkbook(fileName)
        self._cleanTempDir()

    def _setSheetTitle(self, sheet: Worksheet, pcsHeaderPage: PCSHeaderPage, pageCount: int):
        sheet.title = '{}-{}'.format(pcsHeaderPage.processName.replace('/', '-'), pageCount)

    def _mergeHeaders(self, sheet: Worksheet):
        sheet.merge_cells('N3:O3')
        sheet.merge_cells('N4:O4')
        sheet.merge_cells('N5:O5')
        sheet.merge_cells('M7:O7')
        sheet.merge_cells('A9:F9')

    def _mergeBody(self, sheet: Worksheet):
        for i in range(51):
            sheet.merge_cells('E{}:H{}'.format(self.itemStartRow + i, self.itemStartRow + i))

    def _createItemBorder(self, sheet: Worksheet, rowStart: int, pcsItemPage: PCSItemPage):
        sheet.cell(rowStart + pcsItemPage.height - 1, 4).border = Border(
            bottom=Side(style='thin')
        )
        for i in range(7):
            sheet.cell(rowStart + pcsItemPage.height - 1, 5 + i).border = Border(
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='dashed')
            )
        for i in range(4):
            sheet.cell(rowStart + pcsItemPage.height - 1, 12 + i).border = Border(
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin'),
            )
        sheet.cell(rowStart + pcsItemPage.height - 1, 13).border = Border(
            bottom=Side(style='thin'),
        )
        sheet.cell(rowStart + pcsItemPage.height - 1, 14).border = Border(
            bottom=Side(style='thin'),
            right=Side(style='thin'),
        )

        if len(pcsItemPage.controlMethodIntervalType) > 0:
            sheet.cell(rowStart, 9).border = Border(
                bottom=Side(style='dashed'),
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin')
            )
        
        if len(pcsItemPage.controlMethodType) > 0 and pcsItemPage.controlMethodType[0] == 'Auto check':
            sheet.cell(rowStart, 10).border = Border(
                bottom=Side(style='dashed'),
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin')
            )

    def _writeHeaders(self, drawer: PCSSheetDrawer, pcsHeaderPage: PCSHeaderPage):
        #   Draw logo
        drawer.drawDensoLogo()
        #   Write page
        drawer.writeSubHeader(2, 15, 'page {}'.format(pcsHeaderPage.pageCount), 'right', 'bottom')
        #   Write check box
        drawer.writeSubHeader(3, 14, '❑    \t  Prototype', 'left')
        drawer.writeSubHeader(4, 14, '❑    \t  Pre-Launch', 'left')
        drawer.writeSubHeader(5, 14, '❑    \t  Production', 'left')
        #   Write header
        drawer.writeHeader(7, 1, pcsHeaderPage.lineName)
        drawer.writeHeader(7, 8, pcsHeaderPage.assyName)
        drawer.writeHeader(9, 1, pcsHeaderPage.processName, 'left')
        drawer.writeHeader(9, 7, pcsHeaderPage.pageCount)
        drawer.writeHeader(9, 8, pcsHeaderPage.partName)
        drawer.writeHeader(9, 13, pcsHeaderPage.customerName)

        drawer.drawCriticalItems(pcsHeaderPage.criticalItemSummaryDict)
        drawer.writeFooter(63, 9, 'Issue to ❑ Insp.    ❑ Prod.(___________)', 'left', 'center')

    def _saveWorkbook(self, fileName: str):
        templateSheet = self.workbook[self.templateSheetName]
        
        self.workbook.remove(templateSheet)
        self.workbook.save('{outputDir}/{fileName}.xlsx'.format(
            outputDir = self.outputDir,
            fileName = fileName
        ))

    def _createTempDir(self):
        if not os.path.exists(self.tempDir):
            os.mkdir(self.tempDir)

    def _cleanTempDir(self):
        shutil.rmtree(self.tempDir)
        