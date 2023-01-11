from datetime import datetime
from PIL import Image as PILImage, ImageDraw
from openpyxl.styles.fonts import Font
from openpyxl.drawing.image import Image
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.xdr import XDRPositiveSize2D, XDRPoint2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, AbsoluteAnchor
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU, EMU_to_pixels

c2e = cm_to_EMU
p2e = pixels_to_EMU
cellh = lambda x: c2e(x * 0.48)
cellw = lambda x: c2e(x * 1.1)

counterPathMap = {
    1: 'images/counter/1.png',
    2: 'images/counter/2.png',
    3: 'images/counter/3.png',
    4: 'images/counter/4.png',
    5: 'images/counter/5.png',
    6: 'images/counter/6.png',
    7: 'images/counter/7.png',
    8: 'images/counter/8.png',
    9: 'images/counter/9.png',
    10: 'images/counter/10.png',
    11: 'images/counter/11.png',
    12: 'images/counter/12.png',
    13: 'images/counter/13.png',
    14: 'images/counter/14.png',
    15: 'images/counter/15.png',
    16: 'images/counter/16.png',
}
controlItemSymbolPathMap = {
    'None': 'images/timing/check-no-record.png',
    'Check sheet': 'images/timing/check-record.png',
    'Record sheet': 'images/timing/check-record.png',
    'x-R chart': 'images/timing/check-control-chart.png',
    'xbar-R chart': 'images/timing/check-control-chart.png',
    'x-Rs chart': 'images/timing/check-control-chart.png',
}
timingConnectorPath = 'images/timing/check-process.png'
scSymbolPathFormat = 'images/symbols/{}.png'

class PCSSheetDrawer:
    def __init__(self, maxHeight: int, itemStartRow: int, tempDir: str, sheet: Worksheet):
        self.sheet = sheet
        self.tempDir = tempDir
        self.maxHeight = maxHeight
        self.itemStartRow = itemStartRow

    def writeHeader(self, row, col, value, horizontal='center', vertical='center'):
        self.sheet.cell(row=row, column=col).value = value
        self.sheet.cell(row=row, column=col).alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=True
        )
        self.sheet.cell(row=row, column=col).font = Font(name='CordiaUPC', size=14, color='000000', bold=True)

    def writeSubHeader(self, row, col, value, horizontal='center', vertical='center'):
        self.sheet.cell(row=row, column=col).value = value
        self.sheet.cell(row=row, column=col).alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=True
        )
        self.sheet.cell(row=row, column=col).font = Font(name='CordiaUPC', size=12, color='000000')

    def writeFooter(self, row, col, value, horizontal='center', vertical='center'):
        self.sheet.cell(row=row, column=col).value = value
        self.sheet.cell(row=row, column=col).alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
        )
        self.sheet.cell(row=row, column=col).font = Font(name='CordiaUPC', size=14, color='000000')

    def writeBody(self, row, col, valueList, horizontal='center', vertical='center'):
        for i, value in enumerate(valueList):
            self.sheet.cell(row=row + i, column=col).value = value
            self.sheet.cell(row=row + i, column=col).alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
            )
            self.sheet.cell(row=row + i, column=col).font = Font(name='CordiaUPC', size=10, color='000000')

    def writeSubBody(self, row, col, value, horizontal='center', vertical='center'):
        self.sheet.cell(row=row, column=col).value = value
        self.sheet.cell(row=row, column=col).alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
        )
        self.sheet.cell(row=row, column=col).font = Font(name='CordiaUPC', size=12, color='000000')

    def drawCheckTiming(self, pcsItemList: list, summaryTimingDict: dict):
        duringList = summaryTimingDict['During']
        beforeList = summaryTimingDict['Before']
        afterList = summaryTimingDict['After']

        def getRowOfItem(idx: int):
            currentRow = self.itemStartRow
            for i, pcsItemPage in enumerate(pcsItemList):
                if i == idx:
                    break

                currentRow += pcsItemPage.height
            
            return currentRow

        if len(duringList) > 0:
            if len(beforeList) == 0 and len(afterList) == 0:
                landIndex = len(pcsItemList) / 2
                landProcessRow = getRowOfItem(landIndex)
                self.drawCheckProcessConnector(landProcessRow)
            elif len(beforeList) == 0:
                landProcessRow = self.itemStartRow - 1
                self.drawCheckProcessConnector(landProcessRow)
            else:
                landIndex = duringList[int(len(duringList) / 2)]
                landProcessRow = getRowOfItem(landIndex)
                self.drawCheckProcessConnector(landProcessRow)
        else:
            if len(beforeList) == 0:
                landProcessRow = self.itemStartRow - 1
            elif len(afterList) == 0:
                landIndex = len(pcsItemList) - 1
                landProcessRow = getRowOfItem(landIndex)
            else:
                lastBefore = beforeList[-1]
                firstAfter = afterList[0]
                landIndex = (lastBefore + firstAfter) / 2
                landProcessRow = getRowOfItem(landIndex)

        self.drawCheckProcess(landProcessRow)

    def drawScSymbolList(self, rowStart: int, rowHeight: int, scSymbolList: list):
        symbolTotal = len(scSymbolList)
        for i, scSymbol in enumerate(scSymbolList):
            symbolHash = '{}-{}'.format(scSymbol['character'], scSymbol['shape'])
            symbolPath = scSymbolPathFormat.format(symbolHash)
            img = Image(symbolPath)
            if scSymbol['character'] == 'RW':
                rowOff = 0
                colOff = 2.5
                wOff = -5
                hOff = -5
            else:
                rowOff = 0
                colOff = 0
                wOff = 0
                hOff = 0

            if symbolTotal == 1:
                centerOffset = int(rowHeight / 2) - 2 if rowHeight % 2 == 0 else int(rowHeight / 2) - 1
                symbolImg = self._createImage(img, rowStart + centerOffset, 2, -1 + rowOff, 10 + colOff, wOff, hOff)
            else:
                symbolImg = self._createImage(img, rowStart + i -1, 2, 4+(5*i) + rowOff, 10 + colOff, wOff, hOff)
            self.sheet.add_image(symbolImg)

    def drawControlItemSymbol(self, controlItemType: str, rowStart: int, rowHeight: int):
        symbolPath = controlItemSymbolPathMap.get(controlItemType, None)
        if symbolPath is None:
            raise KeyError('Unregistered control item type, {}'.format(controlItemType))

        centerOffset = int(rowHeight / 2) - 2 if rowHeight % 2 == 0 else int(rowHeight / 2) - 1
        img = self._createImage(Image(symbolPath), rowStart + centerOffset, 1, 0, 10)
        self.sheet.add_image(img)
        self._drawHorizontalDashedLine(rowStart + centerOffset, 1, 7, 0)

    def drawControlItemConnectorGroup(self, pcsItemList: list, summaryTimingDict: dict, beforePageItem, afterPageItem):
        def groupControlItem(idxList: list):
            groupDict = dict()
            i = 0
            for idx in idxList:
                if groupDict.get(i, None) is None:
                    groupDict[i] = [idx]
                else:
                    if idx - 1 == groupDict[i][-1]:
                        groupDict[i].append(idx)
                    else:
                        i += 1
                        groupDict[i] = [idx]
            return groupDict

        duringGroupDict = groupControlItem(summaryTimingDict['During'])
        beforeGroupDict = groupControlItem(summaryTimingDict['Before'])
        afterGroupDict = groupControlItem(summaryTimingDict['After'])

        def getRowOfItem(idx: int):
            currentRow = self.itemStartRow
            for i, pcsItemPage in enumerate(pcsItemList):
                if i == idx:
                    break

                currentRow += pcsItemPage.height
            
            return currentRow

        def drawConnectorGroup(idxList: list):
            totalItem = len(idxList)
            if totalItem <= 1:
                return

            totalHeight = 0
            for idx in idxList:
                totalHeight += pcsItemList[idx].height


            self._drawVerticalDashedLine(
                totalHeight - (pcsItemList[0].height / 2) - (pcsItemList[-1].height / 2),
                getRowOfItem(idxList[0]),
                1,
                (pcsItemList[0].height / 2) * 4 + 1,
                0
            )

            if beforePageItem is not None and beforePageItem.checkTiming == pcsItemList[0].checkTiming:
                self._drawVerticalDashedLine(pcsItemList[0].height / 2, self.itemStartRow - 1, 1, 0, 0)

            if afterPageItem is not None and afterPageItem.checkTiming == pcsItemList[-1].checkTiming:
                self._drawVerticalDashedLine(
                    (self.itemStartRow + self.maxHeight) - getRowOfItem(len(pcsItemList) - 1) - 1,
                    getRowOfItem(len(pcsItemList) - 1),
                    1,
                    0,
                    0
                )

        for duringGroupList in duringGroupDict.values():
            drawConnectorGroup(duringGroupList)
        for beforeGroupList in beforeGroupDict.values():
            drawConnectorGroup(beforeGroupList)
        for afterGroupList in afterGroupDict.values():
            drawConnectorGroup(afterGroupList)

    def drawCriticalItems(self, criticalItemCountDict: dict):
        def drawAbsoluteImage(img, rowOff, colOff):
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            position = XDRPoint2D(p2e(523 + colOff), p2e(135 + rowOff))
            img.anchor = AbsoluteAnchor(pos=position, ext=size)
            return img

        i = 0
        for symbolHash, count in criticalItemCountDict.items():
            symbolPath = scSymbolPathFormat.format(symbolHash)

            symbolImg = drawAbsoluteImage(Image(symbolPath), 0, 33*i)
            counterImg = drawAbsoluteImage(Image(counterPathMap[count]), 12, (33 * i) + 23)

            self.sheet.add_image(symbolImg)
            self.sheet.add_image(counterImg)
            i += 1

    def drawDensoLogo(self):
        img = self._createImage(
            Image('images/denso-logo.png'),
            0,
            0,
            0,
            0
        )

        self.sheet.add_image(img)

    def drawCheckProcessDashedLine(self):
        self._drawVerticalDashedLine(self.maxHeight, self.itemStartRow - 1, 0, -0.5, 7.5)

    def drawCheckProcess(self, row):
        img = self._createImage(
            Image(timingConnectorPath),
            row,
            0,
            0,
            0,
        )
        self.sheet.add_image(img)

    def drawCheckProcessConnector(self, row):
        self._drawHorizontalDashedLine(row, 0, 7, 14)

    def _drawHorizontalDashedLine(self, row, col, rowOff, colOff):
        img = self._createImage(
            Image('images/timing/dash-main-to-branch.png'),
            row,
            col,
            rowOff,
            colOff)

        self.sheet.add_image(img)

    def _drawVerticalDashedLine(self, height, row, col, rowOff, colOff):
        height = EMU_to_pixels(c2e(height) * 0.45)
        fileImg = PILImage.new("RGB", (1, height), (255, 255, 255))
        d = ImageDraw.Draw(fileImg)
        cur_y = 0
        space = 4
        length = 4
        for y in range(cur_y, height, length + space):
            d.line([0, y, 0, y + length], fill=(0, 0, 0), width=1)
        filename = '{}/{}.png'.format(self.tempDir, datetime.now().strftime("%Y%m%d%H%M%S%f"))
        fileImg.save(filename)

        img = self._createImage(Image(filename), row, col, rowOff, colOff)
        self.sheet.add_image(img)

    def _createImage(self, img, row, col, rowOff = 0, colOff = 0, wOff = 0, hOff = 0):
        h, w = img.height + hOff, img.width + wOff
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        marker = AnchorMarker(
            row=row,
            col=col,
            rowOff=p2e(rowOff),
            colOff=p2e(colOff)
        )
        img.anchor = OneCellAnchor(marker, size)
        return img

